from django.shortcuts import render
import io
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from django.http import HttpResponse

# data to excel
def export_to_excel(request, items, businessOwner, motortype, location, size, quantity, subtotal_unit, subtotal_totals, materialcost_unit_price, paint_unit_price, nct_unit_price, motor_unit_price, controller_unit_price, fan_unit_price, bellmouth_unit_price, volt_unit_price, jab_unit_price, assembly_unit_price, pack_unit_price, materialcost_total_price, paint_total_price, nct_total_price, motor_total_price, controller_total_price, fan_total_price, bellmouth_total_price, volt_total_price, jab_total_price, assembly_total_price, pack_total_price, transportation_total_price, ffilter_unit_price, ffilter_total_price, direct_unit, direct_totals, maintenance_unit, maintenance_totals, operating_profit_unit, operating_profit_totals, aggregate_unit, aggregate_totals):
    wb = Workbook()
    ws = wb.active

    # 배경색 지정
    colors = {
        "B5:E6": "DDD9C4", 
        "B19:E19": "D8E4BC",  
        "B21:E21": "FFFF00",  
        "B22:E22": "FABF8F",   
        "B25:E25": "C5D9F1"
    }

    for range_string, color in colors.items():
        for row in ws[range_string]:
            for cell in row:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # 폰트 "맑은 고딕"으로 지정
    font = Font(name='맑은 고딕')

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
    
    # 셀 병합 및 병합위치에 내용넣기
    ws.merge_cells('B5:B6')
    ws['B5'] = "품명"
    ws['B5'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('C5:E5')
    ws['C5'] = f"FFU {size}"
    ws['C5'].font = Font(bold=True)
    ws['C5'].alignment = Alignment(horizontal='center', vertical='center')

    # 열 너비 설정
    ws.column_dimensions['A'].width = 1.1
    ws.column_dimensions['B'].width = 26.9
    ws.column_dimensions['C'].width = 6.0
    ws.column_dimensions['D'].width = 11.8
    ws.column_dimensions['E'].width = 15.8

    # 행 높이 설정
    ws.row_dimensions[1].height = 9.6
    ws.row_dimensions[2].height = 30.6
    ws.row_dimensions[3].height = 19.8
    ws.row_dimensions[4].height = 19.8
    ws.row_dimensions[5].height = 30.6
    ws.row_dimensions[6].height = 25.2

    for i in range(7, 27):  
        ws.row_dimensions[i].height = 24.6

    # FFUInput에서 값 불러오기 위해 선언
    businessOwner = businessOwner
    motortype = motortype 
    location = location 
    size = size

    # 현장명, 유형, 지역
    alignment_center_left = Alignment(horizontal="left", vertical="center")
    alignment_center_center = Alignment(horizontal="center", vertical="center")
    alignment_center_right = Alignment(horizontal="right", vertical="center")
    alignment_center = Alignment(vertical="center")

    ws['B2'] = f'▣ 현장명: {businessOwner} FFU'
    ws['B2'].font = Font(bold=True, size = 14)
    ws['B2'].alignment = alignment_center_left

    ws['B3'] = f'▣ 유형: {motortype}'
    ws['B3'].font = Font(size = 14)
    ws['B3'].alignment = alignment_center_left

    ws['B4'] = f'▣ 지역: {location}'
    ws['B4'].font = Font(size = 14)
    ws['B4'].alignment = alignment_center_left

    # 품명, 수량, 단가, 합계의 제목 행
    font_11 = Font(size=11)

    start_row = 6
    cell_to_check = ws.cell(row=start_row, column=2)

    for range_ in ws.merged_cells.ranges:
        if cell_to_check.coordinate in range_:
            cell = ws[range_.start_cell.coordinate]
            cell.value = "품명"
            cell.font = font_11  
            break
    else:
        cell_to_check.value = "품명"
        cell_to_check.font = font_11  

    cell_quantity = ws.cell(row=start_row, column=3, value="수량")
    cell_quantity.font = font_11  

    cell_price = ws.cell(row=start_row, column=4, value="단가")
    cell_price.font = font_11  

    cell_total = ws.cell(row=start_row, column=5, value="합계")
    cell_total.font = font_11

    for cell in ws["C6":"E6"]:
        for c in cell:
            c.alignment = alignment_center_center

    for cell in ws["C7":"C25"]:
        for c in cell:
            c.alignment = alignment_center_center

    for cell in ws["D7":"E25"]:
        for c in cell:
            c.alignment = alignment_center_right
    
    for cell_addr in ["B19", "B21", "B22", "B25", "B26"]:
        ws[cell_addr].alignment = alignment_center_center

    cells_to_center = list(ws["B7":"B18"]) + [(ws["B20"],), (ws["B23"],), (ws["B24"],)]

    for cell_tuple in cells_to_center:
        for cell in cell_tuple:
            cell.alignment = alignment_center


    # 1. 엑셀에서 B7부터 B18까지의 품명
    start_row = 7
    item_names = [
        "자재비 (AL, SPCC 외)", "도장비", "NCT 가공비", "MOTOR", "컨트롤러", "FAN",
        "BELLMOUTH (보호망포함)", "볼트&리벳", "포장용 잡자재", "조립인건비", "포장팔렛트", "운반비"
    ]

    for i, item_name in enumerate(item_names, start=start_row):
        ws.cell(row=i, column=2, value=item_name)

    # 2. C7부터 C17까지, C20, C23부터 C24까지의 셀에 수량(quantity)
    start_row = 7
    quantity_list = [quantity] * 11

    for i, qty in enumerate(quantity_list, start=start_row):
        ws.cell(row=i, column=3, value=format(qty, ","))

    ws.cell(row=20, column=3, value=format(quantity, ","))  # C20
    for i in range(23, 25):
        ws.cell(row=i, column=3, value=format(quantity, ","))  # C23~C24

    # 3. D7부터 D17까지 각 품목에 대한 단가(unit_price)
    unit_prices = [
        format(materialcost_unit_price, ","),
        format(paint_unit_price, ","),
        format(nct_unit_price, ","),
        format(motor_unit_price, ","),
        format(controller_unit_price, ","),
        format(fan_unit_price, ","),
        format(bellmouth_unit_price, ","),
        format(volt_unit_price, ","),
        format(jab_unit_price, ","),
        format(assembly_unit_price, ","),
        format(pack_unit_price, ","),
        format(transportation_total_price, ",")
    ]

    for i, unit_price in enumerate(unit_prices, start=start_row):
        ws.cell(row=i, column=4, value=unit_price)

    # 4. E7부터 E17까지 각 품목에 대한 합계(total_price)
    total_prices = [
        format(materialcost_total_price, ","),
        format(paint_total_price, ","),
        format(nct_total_price, ","),
        format(motor_total_price, ","),
        format(controller_total_price, ","),
        format(fan_total_price, ","),
        format(bellmouth_total_price, ","),
        format(volt_total_price, ","),
        format(jab_total_price, ","),
        format(assembly_total_price, ","),
        format(pack_total_price, ","),
        format(transportation_total_price, ",")
    ]

    for i, total_price in enumerate(total_prices, start=start_row):
        ws.cell(row=i, column=5, value=total_price)

    # UNIT 소계
    ws["B19"].value = "UNIT 소계"
    ws['D19'] = f'\\   {int(subtotal_unit):,}'
    ws['D19'].number_format = "#,##0"

    ws['E19'] = f'\\   {int(subtotal_totals):,}'
    ws['E19'].number_format = "#,##0"

    # FILTER 소계
    ws["B20"].value = "FILTER"
    ws['D20'] = f'{int(ffilter_unit_price):,}'
    ws['E20'] = f'{int(ffilter_total_price):,}'

    ws["B21"].value = "FILTER 소계"
    ws['D21'] = f'\\   {int(ffilter_unit_price):,}'
    ws['E21'] = f'\\   {int(ffilter_total_price):,}'

    # 원가 합계, 간접비, 총계, 비고
    ws["B22"].value = "원가 합계"
    ws['D22'] = f'\\   {int(direct_unit):,}'
    ws['E22'] = f'\\   {int(direct_totals):,}'

    ws["B23"].value = "관리비"
    ws['D23'] = f'{int(maintenance_unit):,}'
    ws['E23'] = f'{int(maintenance_totals):,}'

    ws["B24"].value = "영업이익"
    ws['D24'] = f'{int(operating_profit_unit):,}'
    ws['E24'] = f'{int(operating_profit_totals):,}'

    ws["B25"].value = "총계"
    ws['D25'] = f'\\   {int(aggregate_unit):,}'
    ws['E25'] = f'\\   {int(aggregate_totals):,}'

    ws["B26"].value = "비고"

    ws.merge_cells('C26:E26')
    for cell in ws["C26":"E26"]:
        for c in cell:
            c.alignment = Alignment(horizontal='center')

    # 글씨크기 설정
    font_size_10 = Font(size=10)
    # 각 영역에 대한 글씨 크기 설정
    font_size_ranges = [
        (7, 18, 2, 5, font_size_10),
        (20, 20, 2, 5, font_size_10),
        (23, 24, 2, 5, font_size_10),
        (26, 26, 2, 5, font_size_10),
    ]

    # 각 영역에 대해 설정한 글씨 크기 적용
    for min_row, max_row, min_col, max_col, font in font_size_ranges:
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.font = font

    # 글씨 bold (B19:E19, B21:E22, B25:E25)
    bold_ranges = ['B19:E19', 'B21:E22', 'B25:E25']
    for range_str in bold_ranges:
        for cell in ws[range_str]:
            for c in cell:
                c.font = Font(bold=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    thick_side = Side(style='medium')

    ws = ws 

    # 1. B7부터 E25까지 모든 테두리
    for row in ws.iter_rows(min_row=7, max_row=25, min_col=2, max_col=5):
        for cell in row:
            cell.border = thin_border

    # 2. B26은 오른쪽 테두리
    cell = ws["B26"]
    cell.border = Border(right=thin_border.right)


    # 3. C6부터 E6까지 모든 테두리
    for cell in ws["C6":"E6"]:
        for c in cell:
            c.border = thin_border

    # 4. B5부터 B6까지 오른쪽 테두리
    for cell in ws["B5":"B6"]:
        for c in cell:
            c.border = Border(right=thin_border.right)

    # 5. B5부터 E26까지 굵은 바깥쪽 테두리
    for row in ws.iter_rows(min_row=5, max_row=26, min_col=2, max_col=5):
        for cell in row:
            # 최상단 행이나 최하단 행에 대한 상하측 테두리 설정
            top_border = thick_side if cell.row == 5 else cell.border.top
            bottom_border = thick_side if cell.row == 26 else cell.border.bottom

            # 왼쪽 끝 열이나 오른쪽 끝 열에 대한 좌우측 테두리 설정
            left_border = thick_side if cell.column == 2 else cell.border.left
            right_border = thick_side if cell.column == 5 else cell.border.right

            cell.border = Border(top=top_border, bottom=bottom_border, left=left_border, right=right_border)


    # 6 ~ 10. 굵은 바깥쪽 테두리
    for cell_range in ["B5:E6", "B19:E19", "B21:E21", "B22:E22", "B25:E25"]:
        start_col = ord(cell_range.split(':')[0][0]) - 65
        end_col = ord(cell_range.split(':')[1][0]) - 65
        start_row = int(cell_range.split(':')[0][1:])
        end_row = int(cell_range.split(':')[1][1:])
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col + 1, max_col=end_col + 1):
            for cell in row:
                top_border = thick_side if cell.row == start_row else cell.border.top
                bottom_border = thick_side if cell.row == end_row else cell.border.bottom
                left_border = thick_side if cell.column == start_col + 1 else cell.border.left
                right_border = thick_side if cell.column == end_col + 1 else cell.border.right

                cell.border = Border(top=top_border, bottom=bottom_border, left=left_border, right=right_border)

    return wb

# excel 다운로드
def download_excel(request):
    items = request.session.get('items', [])  
    businessOwner = request.session.get('businessOwner', '')
    motortype = request.session.get('motortype', '')
    location = request.session.get('location', '')
    size = request.session.get('size', 'Unknown Size')
    quantity = request.session.get('quantity', '')
    subtotal_unit = request.session.get('subtotal_unit', '')
    subtotal_totals = request.session.get('subtotal_totals', '')


    # request를 통한 변수 불러오기
    # unit_price (단가)
    materialcost_unit_price = request.session.get('materialcost_unit_price', 0)
    paint_unit_price = request.session.get('paint_unit_price', 0)
    nct_unit_price = request.session.get('nct_unit_price', 0)
    motor_unit_price = request.session.get('motor_unit_price', 0)
    controller_unit_price = request.session.get('controller_unit_price', 0)
    fan_unit_price = request.session.get('fan_unit_price', 0)
    bellmouth_unit_price = request.session.get('bellmouth_unit_price', 0)
    volt_unit_price = request.session.get('volt_unit_price', 0)
    jab_unit_price = request.session.get('jab_unit_price', 0)
    assembly_unit_price = request.session.get('assembly_unit_price', 0)
    pack_unit_price = request.session.get('pack_unit_price', 0)
    ffilter_unit_price = request.session.get('ffilter_unit_price', 0)
    direct_unit = request.session.get('direct_unit', 0)
    maintenance_unit = request.session.get('maintenance_unit', 0)
    operating_profit_unit = request.session.get('operating_profit_unit', 0)
    aggregate_unit = request.session.get('aggregate_unit', 0)    

    # total_price (합계)
    materialcost_total_price = request.session.get('materialcost_total_price', 0)
    paint_total_price = request.session.get('paint_total_price', 0)
    nct_total_price = request.session.get('nct_total_price', 0)
    motor_total_price = request.session.get('motor_total_price', 0)
    controller_total_price = request.session.get('controller_total_price', 0)
    fan_total_price = request.session.get('fan_total_price', 0)
    bellmouth_total_price = request.session.get('bellmouth_total_price', 0)
    volt_total_price = request.session.get('volt_total_price', 0)
    jab_total_price = request.session.get('jab_total_price', 0)
    assembly_total_price = request.session.get('assembly_total_price', 0)
    pack_total_price = request.session.get('pack_total_price', 0)
    transportation_total_price = request.session.get('transportation_total_price', 0)
    ffilter_total_price = request.session.get('ffilter_total_price', 0)
    direct_totals = request.session.get('direct_totals', 0)
    maintenance_totals = request.session.get('maintenance_totals', 0)
    operating_profit_totals = request.session.get('operating_profit_totals', 0)
    aggregate_totals = request.session.get('aggregate_totals', 0)

    wb = export_to_excel(
    request,
    items,
    businessOwner,
    motortype,
    location,
    size,
    quantity,
    subtotal_unit,
    subtotal_totals,
    materialcost_unit_price,
    paint_unit_price,
    nct_unit_price,
    motor_unit_price,
    controller_unit_price,
    fan_unit_price,
    bellmouth_unit_price,
    volt_unit_price,
    jab_unit_price,
    assembly_unit_price,
    pack_unit_price,
    materialcost_total_price,
    paint_total_price,
    nct_total_price,
    motor_total_price,
    controller_total_price,
    fan_total_price,
    bellmouth_total_price,
    volt_total_price,
    jab_total_price,
    assembly_total_price,
    pack_total_price,
    transportation_total_price,
    ffilter_unit_price,
    ffilter_total_price, 
    direct_unit, direct_totals, 
    maintenance_unit, 
    maintenance_totals, 
    operating_profit_unit, 
    operating_profit_totals, 
    aggregate_unit, 
    aggregate_totals
)


    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="FFU.xlsx"'


    return response